from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse_lazy
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, DetailView, View
from django.views import View
from django.contrib.auth import login, logout, authenticate, update_session_auth_hash
from django.contrib.auth.models import User
from django.contrib.auth.forms import UserCreationForm, AuthenticationForm
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import JsonResponse
from django.contrib import messages, auth
from datetime import datetime
from django.db import transaction
from django.utils import timezone
from django.db.models import Count, Q
from .forms import *
from .models import *
import openpyxl
from openpyxl.styles import Font, Alignment
from django.http import HttpResponse
from django.forms import inlineformset_factory, modelformset_factory, ValidationError
from decimal import Decimal 
from django.db.models import Sum, F, Count 
from django.db.models.functions import TruncMonth
from .utils import role_required
from django.template.loader import render_to_string
import json
from django.core.paginator import Paginator

# Fonctions de base pour la gestion des produits
@login_required
@role_required('ADMIN', 'GESTIONNAIRE', 'VENDEUR')
def produit_list(request):
    produits = Produits.objects.all()
    paginator = Paginator(produits, 10)  # 10 produits par page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'produit_list.html', {
        'produits': page_obj,
        'is_paginated': page_obj.has_other_pages()
    })

@login_required
@role_required('ADMIN', 'GESTIONNAIRE', 'VENDEUR')
def produit_detail(request, slug):
    produit = get_object_or_404(Produits, slug=slug)
    return render(request, 'produit_detail.html', {'produit': produit})

@login_required
@role_required('ADMIN', 'GESTIONNAIRE', 'VENDEUR')
def produit_form(request, slug=None):
    instance = None
    if slug:
        instance = get_object_or_404(Produits, slug=slug)

    if request.method == 'POST':
        form = ProduitForm(request.POST, instance=instance)
        if form.is_valid():
            produit = form.save()
            if instance:
                messages.success(request, f"Produit '{produit.name}' modifié avec succès.")
            else:
                messages.success(request, f"Produit '{produit.name}' créé avec succès.")
            return redirect('produit_detail', slug=produit.slug)
    else:
        form = ProduitForm(instance=instance)

    return render(request, 'produit_form.html', {'form': form})



# Generique fonction
class Affichage(LoginRequiredMixin, ListView):
    template_name = 'comptent.html'
    queryset = Stockes.objects.all()
    login_url = 'connexion'





# List view

class detail_vente(LoginRequiredMixin, DetailView):
    model = Vente
    template_name = "detail_vente.html"
    context_object_name = "vente"
    login_url = 'connexion'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['produits_vendus'] = self.object.lignes.all()
        return context


class AddinvoiceView(LoginRequiredMixin, View):
    template_name = "vente.html"
    login_url = 'connexion'

    def get(self, request, *args, **kwargs):
        vente_form = VenteForm()
        customers = Customer.objects.all()
        # Exclure les stocks expirés
        stocks = Stockes.objects.filter(quantite__gt=0).exclude(achat_ligne__date_expiration__lt=timezone.now().date())
        context = {
            'vente_form': vente_form,
            'Customers': customers,
            'Stocks': stocks,
        }
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        vente_form = VenteForm(request.POST)
        if vente_form.is_valid():
            with transaction.atomic():
                produits = request.POST.getlist('stock[]')
                quantites = request.POST.getlist('qt[]')
                prixs = request.POST.getlist('prix[]')

                ligne_erreurs = []
                vente_produits_valides = []
                for i in range(len(produits)):
                    # Vérifier que le stock n'est pas expiré
                    stock_obj = Stockes.objects.get(pk=produits[i])
                    if stock_obj.est_expire:
                        ligne_erreurs.append(f"Ligne {i+1} : Le produit sélectionné est expiré et ne peut pas être vendu.")
                        continue
                    data = {
                        'produit': produits[i],
                        'quantite': quantites[i],
                        'prix_vente':prixs[i]
                    }
                    produit_form = VenteProduitForm(data)
                    if produit_form.is_valid():
                        vente_produits_valides.append(produit_form)
                    else:
                        for field, errors in produit_form.errors.items():
                            for error in errors:
                                if field == '__all__':
                                    ligne_erreurs.append(f"Ligne {i+1} : {error}")
                                else:
                                    ligne_erreurs.append(f"Ligne {i+1} ({field}) : {error}")

                if ligne_erreurs:
                    messages.error(request, "Erreur(s) dans les articles :" + "".join(f"{err}" for err in ligne_erreurs) )
                    customers = Customer.objects.all()
                    stocks = Stockes.objects.filter(quantite__gt=0).exclude(achat_ligne__date_expiration__lt=timezone.now().date())
                    context = {
                        'vente_form': vente_form,
                        'Customers': customers,
                        'Stocks': stocks,
                    }
                    return render(request, self.template_name, context)

                # Si tout est valide, on sauvegarde la vente et les lignes
                vente = vente_form.save(commit=False)
                vente.vendeur = request.user
                vente.save()
                for produit_form in vente_produits_valides:
                    produit_form.instance.vente = vente
                    produit_form.save()

            messages.success(request, "Vente enregistrée avec succès!")
            return redirect('liste_ventes')
        else:
            messages.error(request, "Veuillez corriger les erreurs dans le formulaire principal.")

        customers = Customer.objects.all()
        stocks = Stockes.objects.filter(quantite__gt=0).exclude(achat_ligne__date_expiration__lt=timezone.now().date())
        context = {
            'vente_form': vente_form,
            'Customers': customers,
            'Stocks': stocks,
        }
        return render(request, self.template_name, context)

                

class ListeVentesView(LoginRequiredMixin, ListView):
    model = Vente
    template_name = "liste_ventes.html"
    context_object_name = "ventes"
    login_url = 'connexion'
    ordering = ['-date_vente']
    paginate_by = 10
    
    def get_queryset(self):
        queryset = super().get_queryset().filter(annule=False)
        
        date_debut = self.request.GET.get('date_debut')
        date_fin = self.request.GET.get('date_fin')
        client_id = self.request.GET.get('client')
       
        statut = self.request.GET.get('statut')
        date_payement = self.request.GET.get('date_payement')
        
        if date_debut:
            queryset = queryset.filter(date_vente__gte=date_debut)
        
        if date_fin:
            queryset = queryset.filter(date_vente__lte=date_fin + ' 23:59:59')
        
        if client_id:
            queryset = queryset.filter(customer_id=client_id)

        
        if statut:
            queryset = queryset.filter(statupaiement=statut)
        if date_payement:
            queryset = queryset.filter(date_payement__date=date_payement)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['clients'] = Customer.objects.all()

        # Utiliser une agrégation pour des performances optimales
        full_queryset = self.get_queryset()
        aggregation = full_queryset.aggregate(
            total_ventes=Count('id'),
            montant_total=Sum(F('lignes__quantite') * F('lignes__prix_vente'))
        )

        context['total_ventes'] = aggregation.get('total_ventes', 0)
        context['montant_total'] = aggregation.get('montant_total') or Decimal('0.00')

        return context


@login_required
@role_required('ADMIN', 'GESTIONNAIRE')
def marquer_vente_payee(request, pk):
    """Marquer une vente comme payée."""
    vente = get_object_or_404(Vente, pk=pk)
    
    if request.method == 'POST':
        # Vérifier que la vente est en statut "Dette" avant de la marquer comme "Payée"
        if vente.statupaiement == Vente.PAYMENT_STATUS_DETTE:
            vente.statupaiement = Vente.PAYMENT_STATUS_PAYER  # Statut "Payé"
            vente.date_payement = timezone.now() # Mettre à jour la date de paiement
            vente.save(validate=False) # On bypass la validation car la date de paiement est maintenant dans le passé
            messages.success(request, f"La vente n°{vente.numero_facture} a été marquée comme payée.")
        else:
            messages.warning(request, "Cette vente n'est pas une dette et ne peut être marquée comme payée.")
    else:
        messages.error(request, "Action non autorisée.")
    
    return redirect('liste_ventes')



@login_required
@role_required('ADMIN', 'GESTIONNAIRE')
def liste_utilisateurs(request):
    utilisateurs = CustomUser.objects.all()
    print("Nombre d'utilisateurs:", utilisateurs.count())  # Pour déboguer
    for user in utilisateurs:
        print(user.username, user.email)  # Pour voir les données
    return render(request, 'user/liste_utilisateurs.html', {'utilisateurs': utilisateurs})


@login_required
@role_required('ADMIN', 'GESTIONNAIRE')
def creer_utilisateur(request):
    if request.method == 'POST':
        form = InscriptionForm(request.POST)
        if form.is_valid():
            user = form.save()
            messages.success(request, f"L'utilisateur {user.username} a été créé avec succès.")
            return redirect('liste_utilisateurs')
    else:
        form = InscriptionForm()
    
    return render(request, 'user/inscription.html', {'form': form})

@login_required
@role_required('ADMIN', 'GESTIONNAIRE')
def modifier_utilisateur(request, pk):
    utilisateur = get_object_or_404(CustomUser, pk=pk)
    
    if request.method == 'POST':
        # Utiliser UserEditForm pour la modification
        form = UserEditForm(request.POST, instance=utilisateur)
        if form.is_valid():
            utilisateur_modifie = form.save()
            # Le champ mot de passe n'est pas dans ce formulaire, donc on ne le touche pas.
            messages.success(request, 'Utilisateur modifié avec succès')
            return redirect('liste_utilisateurs')  # Rediriger vers liste plutôt que création
    else:
        form = UserEditForm(instance=utilisateur)
    
    return render(request, 'user/inscription.html', {
        'form': form,
        'utilisateur_courant': utilisateur
    })

@login_required
@role_required('ADMIN', 'GESTIONNAIRE')
def supprimer_utilisateur(request, pk):
    utilisateur = get_object_or_404(CustomUser, pk=pk)
    if request.user == utilisateur:
        messages.error(request, "Vous ne pouvez pas supprimer votre propre compte.")
    elif request.method == 'POST':
            utilisateur.delete()
            messages.success(request, 'Utilisateur supprimé avec succès')
    return redirect('liste_utilisateurs')

def connexion(request):
    if request.method == 'POST':
        form = LoginForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                messages.success(request, 'Connexion réussie!')
                next_url = request.GET.get('next', 'home')
                return redirect(next_url)
            else:
                messages.error(request, 'Nom d\'utilisateur ou mot de passe incorrect.')
    else:
        form = LoginForm()
    return render(request, 'user/connexion.html', {'form': form})

def inscription(request):
    if request.method == 'POST':
        form = InscriptionForm(request.POST)
        if form.is_valid():
            user = form.save()
            messages.success(request, 'Inscription réussie! Vous pouvez maintenant vous connecter.')
            return redirect('connexion')
    else:
        form = InscriptionForm()
    return render(request, 'user/inscription.html', {'form': form})

# Vues pour les catégories
@login_required
@role_required('ADMIN', 'GESTIONNAIRE', 'VENDEUR')
def liste_categories(request):
    categories = Categories.objects.all().order_by('-id')
    paginator = Paginator(categories, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'categorie_list.html', {
        'categories': page_obj,
        'is_paginated': page_obj.has_other_pages()
    })

@login_required
@role_required('ADMIN', 'GESTIONNAIRE', 'VENDEUR')
def ajouter_categorie(request):
    if request.method == 'POST':
        form = CategorieForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Catégorie ajoutée avec succès!')
            return redirect('liste_categories')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"Erreur pour {field}: {error}")
    return redirect('liste_categories')

@login_required
@role_required('ADMIN', 'GESTIONNAIRE', 'VENDEUR')
def modifier_categorie(request, pk):
    categorie = get_object_or_404(Categories, id=pk)
    if request.method == 'POST':
        form = CategorieForm(request.POST, instance=categorie)
        if form.is_valid():
            form.save()
            messages.success(request, 'Catégorie modifiée avec succès!')
            return redirect('liste_categories')
        else:
            # Afficher les erreurs de validation
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"Erreur pour {field}: {error}")
    return redirect('liste_categories')

@login_required
@role_required('ADMIN', 'GESTIONNAIRE', 'VENDEUR')
def supprimer_categorie(request, pk):
    categorie = get_object_or_404(Categories, id=pk)
    if request.method == 'POST':
        try:
            categorie.delete()
            messages.success(request, 'Catégorie supprimée avec succès!')
        except:
            messages.error(request, 'Impossible de supprimer cette catégorie car elle est utilisée par des produits')
    return redirect('liste_categories')

# Vues pour les conditions
@login_required
@role_required('ADMIN', 'GESTIONNAIRE', 'VENDEUR')
def liste_conditions(request):
    conditions = Condition.objects.all().order_by('-id')
    paginator = Paginator(conditions, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'condition_list.html', {
        'conditions': page_obj,
        'is_paginated': page_obj.has_other_pages()
    })

@login_required
@role_required('ADMIN', 'GESTIONNAIRE', 'VENDEUR')
def ajouter_condition(request):
    if request.method == 'POST':
        form = ConditionForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Condition ajoutée avec succès!')
        else:
            for error in form.errors.values():
                messages.error(request, error)
    return redirect('liste_conditions')


@login_required
@role_required('ADMIN', 'GESTIONNAIRE', 'VENDEUR')
def modifier_condition(request, pk):
    condition = get_object_or_404(Condition, id=pk)
    if request.method == 'POST':
        form = ConditionForm(request.POST, instance=condition)
        if form.is_valid():
            form.save()
            messages.success(request, 'Condition modifiée avec succès!')
        else:
            for error in form.errors.values():
                messages.error(request, error)
    return redirect('liste_conditions')

@login_required
@role_required('ADMIN', 'GESTIONNAIRE', 'VENDEUR')
def supprimer_condition(request, pk):
    condition = get_object_or_404(Condition, id=pk)
    if request.method == 'POST':
        try:
            condition.delete()
            messages.success(request, 'Condition supprimée avec succès!')
        except:
            messages.error(request, 'Impossible de supprimer cette condition car elle est utilisée par des stocks')
    return redirect('liste_conditions')

# Vues pour les clients
@login_required
@role_required('ADMIN', 'GESTIONNAIRE', 'VENDEUR')
def liste_customers(request):
    customers = Customer.objects.all().order_by('-id')
    paginator = Paginator(customers, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'customer_list.html', {
        'customers': page_obj,
        'is_paginated': page_obj.has_other_pages()
    })

@login_required
@role_required('ADMIN', 'GESTIONNAIRE', 'VENDEUR')
def ajouter_customer(request):
    if request.method == 'POST':
        form = CustomerForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Client ajouté avec succès!')
        else:
            for error in form.errors.values():
                messages.error(request, error)
    return redirect('liste_customers')

@login_required
@role_required('ADMIN', 'GESTIONNAIRE', 'VENDEUR')
def modifier_customer(request, pk):
    customer = get_object_or_404(Customer, id=pk)
    if request.method == 'POST':
        form = CustomerForm(request.POST, instance=customer)
        if form.is_valid():
            form.save()
            messages.success(request, 'Client modifié avec succès!')
        else:
            for error in form.errors.values():
                messages.error(request, error)
    return redirect('liste_customers')

@login_required
@role_required('ADMIN', 'GESTIONNAIRE', 'VENDEUR')
def supprimer_customer(request, pk):
    customer = get_object_or_404(Customer, id=pk)
    if request.method == 'POST':
        try:
            customer.delete()
            messages.success(request, 'Client supprimé avec succès!')
        except:
            messages.error(request, 'Impossible de supprimer ce client car il est associé à des ventes')
    return redirect('liste_customers')

@login_required
@role_required('ADMIN', 'GESTIONNAIRE', 'VENDEUR')
def export_ventes_excel(request):
    # Récupérer les filtres depuis la requête
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    client_id = request.GET.get('client')
    statut = request.GET.get('statut')

    # Filtrer les ventes
    ventes = Vente.objects.all()
    if date_debut:
        ventes = ventes.filter(date_vente__gte=date_debut)
    if date_fin:
        ventes = ventes.filter(date_vente__lte=date_fin + ' 23:59:59')
    if client_id:
        ventes = ventes.filter(customer_id=client_id)
    if statut:
        ventes = ventes.filter(statupaiement=statut)

    if not ventes.exists():
        messages.error(request, "Aucune vente trouvée pour les critères sélectionnés.")
        return redirect('liste_ventes')

    # Créer un fichier Excel
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Rapport des Ventes"

    headers = ["ID Vente", "Client", "statut", "Date de Vente", "Date de payement", "Produit", "Quantité", "Prix Total"]
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    row_num = 2
    for vente in ventes:
        for produit_vendu in vente.lignes.all():  # <-- ici, on utilise 'lignes'
            sheet.cell(row=row_num, column=1).value = vente.id
            sheet.cell(row=row_num, column=2).value = vente.customer.name
            sheet.cell(row=row_num, column=3).value = vente.statupaiement
            sheet.cell(row=row_num, column=4).value = vente.date_vente.strftime('%Y-%m-%d')
            sheet.cell(row=row_num, column=5).value = vente.date_payement.strftime('%Y-%m-%d') if vente.date_payement else ""
            sheet.cell(row=row_num, column=6).value = produit_vendu.produit.produit.name
            sheet.cell(row=row_num, column=7).value = produit_vendu.quantite
            sheet.cell(row=row_num, column=8).value = produit_vendu.quantite * produit_vendu.prix_vente
            row_num += 1

    # Ajuster la largeur des colonnes
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        sheet.column_dimensions[column_letter].width = max_length + 2

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="rapport_ventes.xlsx"'
    workbook.save(response)
    return response


@login_required
@role_required('ADMIN', 'GESTIONNAIRE', 'VENDEUR')
def ajuster_stock(request):
    if request.method == 'POST':
        form = AjustementStockForm(request.POST)
        if form.is_valid():
            with transaction.atomic():
                stock_initial = form.cleaned_data['stock']
                stock = Stockes.objects.select_for_update().get(pk=stock_initial.pk)

                motif = form.cleaned_data['motif']
                quantite_ajustee = form.cleaned_data['quantite']
                raison_details = form.cleaned_data['raison']

                quantite_avant = stock.quantite

                # Déterminer le type de modification pour un meilleur suivi
                if motif in ['RETOUR_CLIENT', 'AJUSTEMENT_POSITIF']:
                    type_modif = 'ENTREE'
                    stock.quantite += quantite_ajustee
                    message_action = "ajoutée au"
                else: # SORTIE
                    type_modif = 'SORTIE'
                    if quantite_ajustee > stock.quantite:
                        messages.error(request, f"La quantité à retirer ({quantite_ajustee}) ne peut pas être supérieure au stock actuel ({stock.quantite}).")
                        return redirect('ajuster_stock')
                    stock.quantite -= quantite_ajustee
                    message_action = "retirée du"

                quantite_apres = stock.quantite

                # Mise à jour du stock et traçabilité
                stock.save()
                ModificationStock.objects.create(
                    produit=stock,
                    type_modification=type_modif,
                    motif=motif,
                    quantite=quantite_ajustee,
                    utilisateur=request.user,
                    raison=raison_details,
                    quantite_avant=quantite_avant,
                    quantite_apres=quantite_apres
                )
                messages.success(request, f"Ajustement réussi. {quantite_ajustee} unité(s) a/ont été {message_action} stock pour '{stock.produit.name}'.")
                return redirect('liste_stock')

    else:
        form = AjustementStockForm()
    return render(request, 'ajuster_stock.html', {'form': form})

@login_required
@role_required('ADMIN', 'GESTIONNAIRE', 'VENDEUR')
def ajouter_achat(request):
    AchatLigneFormSet = modelformset_factory(AchatLigne, form=AchatLigneForm, extra=3)
    if request.method == 'POST':
        achat_form = AchatForm(request.POST)
        formset = AchatLigneFormSet(request.POST, queryset=AchatLigne.objects.none())
        if achat_form.is_valid() and formset.is_valid():
            try:
                with transaction.atomic():  # Utiliser une transaction pour garantir l'intégrité des données
                    # Sauvegarder l'achat
                    achat = achat_form.save(commit=False)
                    achat.utilisateur = request.user
                    achat.save()

                    # Traiter les lignes d'achat
                    for form in formset:
                        if form.cleaned_data:  # Vérifier qu'il y a des données
                            # Créer la ligne d'achat
                            ligne = form.save(commit=False)
                            ligne.achat = achat
                            ligne.save()

                            quantite_avant = 0
                            quantite_apres = ligne.quantite

                            # Créer une entrée dans le stock
                            stock = Stockes.objects.create(
                                produit=ligne.produit,
                                quantite=ligne.quantite,
                                prix_achat=ligne.prix_achat , #* Decimal('1.2'),  # Marge de 20%
                                stock_minimum=5,
                                achat_ligne=ligne
                            )

                            # Enregistrer la modification de stock
                            ModificationStock.objects.create(
                                produit=stock,
                                type_modification='ENTREE',
                                motif='ACHAT',
                                quantite=ligne.quantite,
                                utilisateur=request.user,
                                raison=f"Achat n°{achat.id} - Lot: {ligne.lot}",
                                quantite_avant=quantite_avant,
                                quantite_apres=quantite_apres
                            )

                messages.success(request, "Achat enregistré et stock mis à jour avec succès !")
                return redirect('liste_achats')

            except Exception as e:
                messages.error(request, f"Erreur lors de l'enregistrement : {str(e)}")
        else:
            # Les erreurs de formulaire seront affichées dans le template
            messages.error(request, "Veuillez corriger les erreurs ci-dessous.")
    else:
        achat_form = AchatForm()
        formset = AchatLigneFormSet(queryset=AchatLigne.objects.none())

    return render(request, 'ajouter_achat.html', {
        'achat_form': achat_form,
        'formset': formset
    })


class ListeStockView(LoginRequiredMixin, ListView):
    model = Stockes
    template_name = 'liste_stock.html'
    context_object_name = 'stocks'
    login_url = 'connexion'
    ordering = ['-achat_ligne__date_ajout']  # Correction: utiliser le champ via la relation
    paginate_by = 10  # Optionnel, pour paginer

    def get_queryset(self):
        queryset = Stockes.objects.select_related('produit', 'achat_ligne__conditionnement').all()
        
        statut_exp = self.request.GET.get('statut_exp', '')
        statut_q = self.request.GET.get('statut_q')
        produit_id = self.request.GET.get('produit')
        today = timezone.now().date()

        # Filtrage par statut d'expiration
        if statut_exp == 'expire':
            queryset = queryset.filter(achat_ligne__date_expiration__lt=today)
        elif statut_exp == 'valide':
            queryset = queryset.filter(Q(achat_ligne__date_expiration__gte=today) | Q(achat_ligne__date_expiration__isnull=True))
        elif not statut_exp:  # Comportement par défaut : afficher les produits valides
            queryset = queryset.filter(Q(achat_ligne__date_expiration__gte=today) | Q(achat_ligne__date_expiration__isnull=True))

        # Filtrage par statut de quantité (en stock, limité, rupture)
        if statut_q == 'green':
            queryset = queryset.filter(quantite__gt=F('stock_minimum'))
        elif statut_q == 'orange':
            queryset = queryset.filter(quantite__lte=F('stock_minimum'), quantite__gt=0)
        elif statut_q == 'red':
            queryset = queryset.filter(quantite=0)

        # Filtrage par produit si demandé
        if produit_id:
            queryset = queryset.filter(produit_id=produit_id)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        all_stocks = Stockes.objects.all()
        today = timezone.now().date()
        stocks_disponibles = all_stocks.filter(Q(achat_ligne__date_expiration__gte=today) | Q(achat_ligne__date_expiration__isnull=True))
        
        context['nb_en_stock_disponible'] = stocks_disponibles.filter(quantite__gt=F('stock_minimum')).count()
        context['nb_limite_disponible'] = stocks_disponibles.filter(quantite__lte=F('stock_minimum'), quantite__gt=0).count()
        context['nb_rupture'] = all_stocks.filter(quantite=0).count()
        context['nb_expire'] = all_stocks.filter(achat_ligne__date_expiration__lt=today).count()
        context['total_produits'] = all_stocks.count()
        context['filtre_statut_exp'] = self.request.GET.get('statut_exp', '')
        context['filtre_produit'] = self.request.GET.get('produit', '')
        context['produits'] = Produits.objects.all()
        return context

class DetailStockView(LoginRequiredMixin, DetailView):
    model = Stockes
    template_name = 'detail_stock.html'
    context_object_name = 'stock'
    login_url = 'connexion'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Ajoute ici l'historique des mouvements si besoin
        context['modifications'] = self.object.modifications.all().order_by('-date_modification')
        return context

class ListeAchatsView(LoginRequiredMixin, ListView):
    
    model = Achat
    template_name = 'liste_achats.html'
    context_object_name = 'achats'
    login_url = 'connexion'
    ordering = ['-date_achat']
    paginate_by = 10
    
    def get_queryset(self):
        return Achat.objects.all().order_by('-date_achat')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class DetailAchatView(LoginRequiredMixin, DetailView):
    model = Achat
    template_name = 'detail_achat.html'
    context_object_name = 'achat'
    login_url = 'connexion'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['lignes'] = self.object.lignes.all()
        return context

class AchatLigneCreateView(LoginRequiredMixin, CreateView):
    model = AchatLigne
    form_class = AchatLigneForm
    template_name = 'ajouter_achat.html'
    success_url = reverse_lazy('liste_achats')

    def form_valid(self, form):
        with transaction.atomic():
            achat_ligne = form.save()

            quantite_avant = 0
            quantite_apres = achat_ligne.quantite

            # Créer automatiquement une entrée dans Stockes
            stock = Stockes.objects.create(
                produit=achat_ligne.produit,
                quantite=achat_ligne.quantite,
                prix_achat=achat_ligne.prix_achat * Decimal('1.2'),  # Marge de 20% par défaut
                stock_minimum=5,  # Valeur par défaut
                achat_ligne=achat_ligne
            )
            # Créer une modification de stock
            ModificationStock.objects.create(
                produit=stock,
                type_modification='ENTREE',
                quantite=achat_ligne.quantite,
                utilisateur=self.request.user,
                raison=f"Nouvel achat - Lot: {achat_ligne.lot}",
                quantite_avant=quantite_avant,
                quantite_apres=quantite_apres
            )
        return super().form_valid(form)

@login_required
@role_required('ADMIN', 'GESTIONNAIRE', 'VENDEUR')
def annuler_vente(request, pk):
    vente = get_object_or_404(Vente, pk=pk)
    if not vente.annule:
        with transaction.atomic():
            vente.annule = True
            vente.save(validate=False)  # Désactive la validation métier
            # Réajuster le stock pour chaque ligne de vente
            for ligne in vente.lignes.all():
                stock = Stockes.objects.select_for_update().get(pk=ligne.produit.pk)

                quantite_avant = stock.quantite
                stock.quantite += ligne.quantite
                quantite_apres = stock.quantite

                stock.save()
                ModificationStock.objects.create(
                    produit=stock,
                    type_modification='ENTREE',
                    motif='ANNULATION_VENTE',
                    quantite=ligne.quantite,
                    utilisateur=request.user,
                    raison=f"Annulation vente n°{vente.numero_facture}",
                    quantite_avant=quantite_avant,
                    quantite_apres=quantite_apres
                )
            messages.success(request, f"Vente n°{vente.id} annulée et stock réajusté.")
    else:
        messages.info(request, "Cette vente est déjà annulée.")
    return redirect('liste_ventes')

@login_required
@role_required('ADMIN', 'GESTIONNAIRE', 'VENDEUR')
def annuler_achat(request, pk):
    achat = get_object_or_404(Achat, pk=pk)
    if not achat.annule:
        with transaction.atomic():
            achat.annule = True
            achat.save()
            # Réajuster le stock pour chaque ligne d'achat
            for ligne in achat.lignes.all():
                stocks = Stockes.objects.filter(achat_ligne=ligne)
                for stock_initial in stocks:
                    stock = Stockes.objects.select_for_update().get(pk=stock_initial.pk)

                    quantite_avant = stock.quantite
                    stock.quantite -= ligne.quantite
                    quantite_apres = stock.quantite

                    stock.save()
                    ModificationStock.objects.create(
                        produit=stock,
                        type_modification='SORTIE',
                        motif='ANNULATION_ACHAT',
                        quantite=ligne.quantite,
                        utilisateur=request.user,
                        raison=f"Annulation achat n°{achat.id}",
                        quantite_avant=quantite_avant,
                        quantite_apres=quantite_apres
                    )
            messages.success(request, f"Achat n°{achat.id} annulé et stock réajusté.")
    else:
        messages.info(request, "Cet achat est déjà annulé.")
    return redirect('liste_achats')

@login_required
@role_required('ADMIN', 'GESTIONNAIRE', 'VENDEUR')
def liste_fournisseurs(request):
    fournisseurs = Fournisseur.objects.all().order_by('-id')
    paginator = Paginator(fournisseurs, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'liste_fournisseurs.html', {
        'fournisseurs': page_obj,
        'is_paginated': page_obj.has_other_pages()
    })

@login_required
@role_required('ADMIN', 'GESTIONNAIRE', 'VENDEUR')
def ajouter_fournisseur(request):
    if request.method == 'POST':
        form = FournisseurForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Fournisseur ajouté avec succès!')
            return redirect('liste_fournisseurs')
    else:
        form = FournisseurForm()
    return render(request, 'ajout_fourniseur.html', {'form': form})

@login_required
@role_required('ADMIN', 'GESTIONNAIRE', 'VENDEUR')
def modifier_fournisseur(request, pk):
    fournisseur = get_object_or_404(Fournisseur, pk=pk)
    if request.method == 'POST':
        form = FournisseurForm(request.POST, instance=fournisseur)
        if form.is_valid():
            form.save()
            messages.success(request, 'Fournisseur modifié avec succès!')
            return redirect('liste_fournisseurs')
    else:
        form = FournisseurForm(instance=fournisseur)
    return render(request, 'ajout_fourniseur.html', {'form': form, 'fournisseur': fournisseur})

@login_required
@role_required('ADMIN', 'GESTIONNAIRE', 'VENDEUR')
def supprimer_fournisseur(request, pk):
    fournisseur = get_object_or_404(Fournisseur, pk=pk)
    if request.method == 'POST':
        try:
            fournisseur.delete()
            messages.success(request, 'Fournisseur supprimé avec succès!')
        except Exception as e:
            messages.error(request, 'Impossible de supprimer ce fournisseur car il est lié à des achats.')
    return redirect('liste_fournisseurs')

class DashboardView(LoginRequiredMixin, View):
    template_name = 'dashboard.html'

    def get(self, request):
        today = timezone.now().date()
        current_month = today.month
        current_year = today.year

        # --- Prix d'achat total du mois (tous les achats, pas seulement vendus) ---
        total_achats_mois = AchatLigne.objects.filter(
            achat__date_achat__year=current_year,
            achat__date_achat__month=current_month,
            achat__annule=False
        ).aggregate(
            total=Sum(F('prix_achat') * F('quantite'))
        )['total'] or 0
        today = timezone.now().date()
        current_month = today.month
        current_year = today.year
        
        # --- Statistiques des stocks (requêtes optimisées) ---
        all_stocks = Stockes.objects.all()
        
        stock_faible = all_stocks.filter(
            quantite__lte=F('stock_minimum'), quantite__gt=0,
            achat_ligne__date_expiration__gte=today
        ).count()
        
        produits_expires = all_stocks.filter(achat_ligne__date_expiration__lt=today).count()
        
        date_limite_30_jours = today + timezone.timedelta(days=30)
        produits_proche_expiration = all_stocks.filter(
            achat_ligne__date_expiration__gte=today,
            achat_ligne__date_expiration__lte=date_limite_30_jours
        ).count()

        # --- Statistiques des ventes (agrégation unique) ---
        ventes_mois_agg = Vente.objects.filter(
            date_vente__year=current_year,
            date_vente__month=current_month,
            annule=False
        ).aggregate(
            total_ventes=Sum(F('lignes__quantite') * F('lignes__prix_vente')),
            nb_ventes=Count('id')
        )
        total_ventes_mois = ventes_mois_agg['total_ventes'] or 0
        nb_ventes_mois = ventes_mois_agg['nb_ventes'] or 0

        # --- Top 5 des produits les plus vendus ---
        top_produits = VenteProduit.objects.values(
            'produit__produit__name'
        ).annotate(
            total_vendu=Sum('quantite')
        ).order_by('-total_vendu')[:5]
        
        # --- Coût d'achat des produits vendus ce mois ---
        cout_achat_vendu = VenteProduit.objects.filter(
            vente__date_vente__year=current_year,
            vente__date_vente__month=current_month,
            vente__annule=False
        ).aggregate(
            total=Sum(F('quantite') * F('produit__prix_achat'))
        )['total'] or 0

        # --- Marge brute réelle ---
        marge_brute = total_ventes_mois - cout_achat_vendu

        # --- Données pour le graphique des ventes ---
        twelve_months_ago = today - timezone.timedelta(days=365)
        sales_data = Vente.objects.filter(
            date_vente__gte=twelve_months_ago,
            annule=False
        ).annotate(
            month=TruncMonth('date_vente')
        ).values('month').annotate(
            total_sales=Sum(F('lignes__quantite') * F('lignes__prix_vente'))
        ).order_by('month')

        # Préparer les données pour Chart.js
        sales_labels = [s['month'].strftime('%b %Y') for s in sales_data]
        sales_values = [float(s['total_sales'] or 0) for s in sales_data]
        
        context = {
            'total_produits': all_stocks.count(),
            'stock_faible': stock_faible,
            'produits_expires': produits_expires,
            'produits_proche_expiration': produits_proche_expiration,
            'ventes_mois': nb_ventes_mois,
            'total_ventes_mois': total_ventes_mois,
            'top_produits': top_produits,
            'cout_achat_vendu': cout_achat_vendu,
            'total_achats_mois': total_achats_mois,
            'marge_brute': marge_brute,
            'sales_labels': json.dumps(sales_labels),
            'sales_values': json.dumps(sales_values),
        }
        
        return render(request, self.template_name, context)

@login_required
@role_required('ADMIN', 'GESTIONNAIRE', 'VENDEUR')
def facture_vente_view(request, pk):
    vente = get_object_or_404(Vente.objects.prefetch_related('lignes__produit__produit'), pk=pk)
    entreprise = Entreprise.objects.first()
    phrase = "Merci de votre achat !"
    context = {
        'vente': vente,
        'entreprise': entreprise,
        'phrase': phrase,
    }
    return render(request, 'facture_vente.html', context)

@login_required
@role_required('ADMIN', 'GESTIONNAIRE', 'VENDEUR')
def historique_ventes(request):
    """
    Affiche un historique des produits vendus avec filtres et pagination.
    """
    historique = VenteProduit.objects.filter(vente__annule=False).select_related(
        'produit__produit', 
        'vente',
        'modification'
    ).order_by('-vente__date_vente')

    # Filtres
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    produit_id = request.GET.get('produit')

    if date_debut:
        historique = historique.filter(vente__date_vente__gte=date_debut)
    if date_fin:
        historique = historique.filter(vente__date_vente__lte=date_fin + ' 23:59:59')
    if produit_id:
        historique = historique.filter(produit__produit_id=produit_id)

    # Ajout de la pagination
    paginator = Paginator(historique, 10)  # 10 articles par page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'historique': page_obj,
        'is_paginated': page_obj.has_other_pages(),
        'produits': Produits.objects.all()
    }
    return render(request, 'historique_ventes.html', context)
